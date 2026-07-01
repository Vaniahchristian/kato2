"""Parse Super Will Limited shipping manifest CSV/XLSX files."""

from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import BinaryIO


@dataclass
class PricingLineItem:
    shop_no: str
    description: str
    ctns: int | None = None
    qty_per_ctn: int | None = None
    unit_price: float | None = None
    unit_cbm: float | None = None
    yen_ugx: float | None = None
    exp_per_ctn: float | None = None
    value_nature: str | None = None
    value_amount: float | None = None
    value_px: float | None = None
    total_exp_per_ctn: float | None = None
    selling_px_per_ctn: float | None = None
    prt_per_ctn: float | None = None
    tt_pft_per_ctn: float | None = None
    tt_sales: float | None = None
    tt_taxes: float | None = None


@dataclass
class ParsedPricingSheet:
    sheet_name: str
    client: str | None = None
    items: list[PricingLineItem] = field(default_factory=list)
    general_cartons: int | None = None
    general_cbm: float | None = None
    general_weight: float | None = None


@dataclass
class ParseResult:
    manifest: "ParsedInvoice"
    pricing: ParsedPricingSheet | None = None


@dataclass
class LineItem:
    description: str
    ctns: int | None
    qty_per_ctn: int | None
    unit: str | None
    unit_price: float | None
    price_suffix: str | None
    amount: float | None
    weight_kg: float | None
    gross_weight_kg: float | None


@dataclass
class ShopGroup:
    shop_no: str
    sub_amount: float | None = None
    deposit: float | None = None
    balance: float | None = None
    cbm: float | None = None
    line_items: list[LineItem] = field(default_factory=list)


@dataclass
class ParsedInvoice:
    supplier_name: str | None = None
    client: str | None = None
    loading_date: date | None = None
    container_no: str | None = None
    shop_groups: list[ShopGroup] = field(default_factory=list)
    general_cartons: int | None = None
    general_amount: float | None = None
    general_cbm: float | None = None
    general_weight: float | None = None
    fx_rate: float | None = None
    usd_amount: float | None = None
    deposit: float | None = None
    commission: float | None = None
    ship_cost: float | None = None
    balance: float | None = None


class ParseError(Exception):
    def __init__(self, message: str, row: int | None = None):
        self.row = row
        super().__init__(message)


def _cell_val(c) -> str:
    if c is None:
        return ""
    if isinstance(c, datetime):
        return c.strftime("%Y-%m-%d")
    if isinstance(c, date):
        return c.isoformat()
    return str(c).strip()


def _row_to_strings(row) -> list[str]:
    return [_cell_val(c) for c in row]


def _is_pricing_sheet(rows: list[list[str]]) -> bool:
    for row in rows:
        if _cell(row, 0) == "Client" and _cell(row, 5) == "UNIT CBM":
            return True
    return False


def _is_manifest_sheet(rows: list[list[str]]) -> bool:
    for row in rows:
        if _cell(row, 0) == "Client" and "Loading Date" in _cell(row, 2):
            return True
    return False


def _cell(row: list[str], idx: int) -> str:
    if idx < len(row):
        return (row[idx] or "").strip()
    return ""


def _to_float(val: str) -> float | None:
    val = val.strip().replace(",", "").replace("￥", "").replace("$", "")
    if not val:
        return None
    try:
        return float(val)
    except ValueError:
        return None


def _to_int(val: str) -> int | None:
    f = _to_float(val)
    return int(f) if f is not None else None


def _parse_date(val: str) -> date | None:
    val = val.strip()
    if not val:
        return None
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(val, fmt).date()
        except ValueError:
            continue
    return None


def _read_xlsx_workbook(data: bytes) -> tuple[list[list[str]], ParsedPricingSheet | None]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    manifest_rows: list[list[str]] = []
    pricing: ParsedPricingSheet | None = None
    for ws in wb.worksheets:
        sheet_rows = [_row_to_strings(row) for row in ws.iter_rows(values_only=True)]
        if _is_manifest_sheet(sheet_rows):
            manifest_rows.extend(sheet_rows)
        elif _is_pricing_sheet(sheet_rows):
            pricing = parse_pricing_rows(sheet_rows, ws.title)
    wb.close()
    if not manifest_rows:
        raise ParseError("No manifest sheet found (expected Client + Loading Date header)")
    return manifest_rows, pricing


def _read_rows_from_bytes(data: bytes, filename: str) -> tuple[list[list[str]], ParsedPricingSheet | None]:
    lower = filename.lower()
    if lower.endswith(".xlsx"):
        return _read_xlsx_workbook(data)

    text = data.decode("utf-8-sig")
    return list(csv.reader(io.StringIO(text))), None


def _read_rows_from_path(path: Path) -> tuple[list[list[str]], ParsedPricingSheet | None]:
    return _read_rows_from_bytes(path.read_bytes(), path.name)


def _pricing_tt_sales(row: list[str]) -> float | None:
    return _to_float(_cell(row, 15)) or _to_float(_cell(row, 17))


def _pricing_tt_taxes(row: list[str]) -> float | None:
    return _to_float(_cell(row, 16)) or _to_float(_cell(row, 18))


def parse_pricing_rows(rows: list[list[str]], sheet_name: str) -> ParsedPricingSheet:
    sheet = ParsedPricingSheet(sheet_name=sheet_name)
    current_shop = ""

    for i, row in enumerate(rows, start=1):
        c0 = _cell(row, 0)
        if not c0 and not _cell(row, 1):
            continue

        if c0 and "Limited" in c0:
            continue

        if c0 == "Client":
            sheet.client = _cell(row, 1) or sheet.client
            continue

        if c0 in ("Shop No.", "Ctns"):
            continue

        if c0 == "GENERAL CARTONS":
            sheet.general_cartons = _to_int(_cell(row, 2))
            sheet.general_cbm = _to_float(_cell(row, 13))
            sheet.general_weight = _to_float(_cell(row, 16))
            continue

        if c0.upper().startswith(("DEPOSIT", "COMMISSION", "SHIP", "BALANCE")) or "￥" in c0 or "¥" in c0:
            continue

        desc = _cell(row, 1)
        if not desc:
            continue

        if c0:
            current_shop = c0

        if not current_shop:
            raise ParseError(f"Pricing line without shop: {desc}", row=i)

        sheet.items.append(
            PricingLineItem(
                shop_no=current_shop,
                description=desc,
                ctns=_to_int(_cell(row, 2)),
                qty_per_ctn=_to_int(_cell(row, 3)),
                unit_price=_to_float(_cell(row, 4)),
                unit_cbm=_to_float(_cell(row, 5)),
                yen_ugx=_to_float(_cell(row, 6)),
                exp_per_ctn=_to_float(_cell(row, 7)),
                value_nature=_cell(row, 8) or None,
                value_amount=_to_float(_cell(row, 9)),
                value_px=_to_float(_cell(row, 10)),
                total_exp_per_ctn=_to_float(_cell(row, 11)),
                selling_px_per_ctn=_to_float(_cell(row, 12)),
                prt_per_ctn=_to_float(_cell(row, 13)),
                tt_pft_per_ctn=_to_float(_cell(row, 14)),
                tt_sales=_pricing_tt_sales(row),
                tt_taxes=_pricing_tt_taxes(row),
            )
        )

    if not sheet.items:
        raise ParseError(f"No pricing items found in {sheet_name}")

    return sheet


def _parse_footer_line(line: str, invoice: ParsedInvoice) -> None:
    # FX: ￥62101.4/7.15=$8685.510
    fx = re.search(r"[/=]([\d.]+)\s*=\s*\$?([\d.]+)", line)
    if "￥" in line or "¥" in line:
        rate = re.search(r"[/]([\d.]+)", line)
        if rate:
            invoice.fx_rate = float(rate.group(1))
        if fx:
            invoice.usd_amount = float(fx.group(2))

    if line.upper().startswith("DEPOSIT"):
        m = re.search(r"\$?([\d.]+)", line)
        if m:
            invoice.deposit = float(m.group(1))

    if line.upper().startswith("COMMISSION"):
        amounts = re.findall(r"\$?([\d.]+)", line)
        if amounts:
            invoice.commission = float(amounts[-1])

    if line.upper().startswith("SHIP"):
        amounts = re.findall(r"=\$?([\d.]+)", line)
        if amounts:
            invoice.ship_cost = float(amounts[-1])
        else:
            m = re.search(r"\$?([\d.]+)\s*$", line)
            if m:
                invoice.ship_cost = float(m.group(1))

    if line.upper().startswith("BALANCE"):
        amounts = re.findall(r"=\$?([\d.]+)", line)
        if amounts:
            invoice.balance = float(amounts[-1])


def parse_rows(rows: list[list[str]]) -> ParsedInvoice:
    invoice = ParsedInvoice()
    current_group: ShopGroup | None = None
    in_items = False

    for i, row in enumerate(rows, start=1):
        c0 = _cell(row, 0)
        if not c0 and not any(_cell(row, j) for j in range(1, 14)):
            continue

        # Supplier
        if c0 and c0 != "Client" and not in_items and not invoice.supplier_name:
            if "Limited" in c0 or "LTD" in c0.upper():
                invoice.supplier_name = c0
                continue

        if c0 == "Client":
            invoice.client = _cell(row, 1) or invoice.client
            invoice.loading_date = _parse_date(_cell(row, 5)) or invoice.loading_date
            invoice.container_no = _cell(row, 11) or invoice.container_no
            continue

        if c0 == "Shop No.":
            in_items = True
            continue

        if c0 == "GENERAL CARTONS":
            in_items = False
            invoice.general_cartons = _to_int(_cell(row, 2))
            invoice.general_amount = _to_float(_cell(row, 7))
            invoice.general_cbm = _to_float(_cell(row, 10))
            invoice.general_weight = _to_float(_cell(row, 13))
            continue

        upper = c0.upper()
        if upper.startswith("DEPOSIT") or upper.startswith("COMMISSION") or upper.startswith("SHIP") or upper.startswith("BALANCE") or "￥" in c0 or "¥" in c0:
            _parse_footer_line(c0, invoice)
            continue

        if not in_items:
            continue

        desc = _cell(row, 1)
        if not desc:
            continue

        if c0:
            current_group = ShopGroup(
                shop_no=c0,
                sub_amount=_to_float(_cell(row, 8)),
                deposit=_to_float(_cell(row, 9)),
                balance=_to_float(_cell(row, 10)),
                cbm=_to_float(_cell(row, 11)),
            )
            invoice.shop_groups.append(current_group)
        elif current_group is None:
            raise ParseError(f"Line item without shop group: {desc}", row=i)

        current_group.line_items.append(
            LineItem(
                description=desc,
                ctns=_to_int(_cell(row, 2)),
                qty_per_ctn=_to_int(_cell(row, 3)),
                unit=_cell(row, 4) or None,
                unit_price=_to_float(_cell(row, 5)),
                price_suffix=_cell(row, 6) or None,
                amount=_to_float(_cell(row, 7)),
                weight_kg=_to_float(_cell(row, 12)),
                gross_weight_kg=_to_float(_cell(row, 13)),
            )
        )

    if not invoice.shop_groups:
        raise ParseError("No shop groups found in file")

    return invoice


def parse_file(path: str | Path) -> ParseResult:
    rows, pricing = _read_rows_from_path(Path(path))
    return ParseResult(manifest=parse_rows(rows), pricing=pricing)


def parse_upload(data: bytes, filename: str) -> ParseResult:
    rows, pricing = _read_rows_from_bytes(data, filename)
    return ParseResult(manifest=parse_rows(rows), pricing=pricing)


def _self_check() -> None:
    root = Path(__file__).resolve().parent.parent
    for name in ("NALUYIMA 04 (1).csv", "NALUYIMA 04 (1).xlsx"):
        sample = root / name
        if not sample.exists():
            continue
        inv = parse_file(sample).manifest
        assert inv.supplier_name == "Super Will Limited", (name, inv.supplier_name)
        assert inv.client == "NALUYIMA", (name, inv.client)
        assert inv.container_no == "TGBU9257280", (name, inv.container_no)
        assert inv.loading_date == date(2025, 4, 15), (name, inv.loading_date)
        assert len(inv.shop_groups) == 7, (name, len(inv.shop_groups))
        total_items = sum(len(g.line_items) for g in inv.shop_groups)
        assert total_items == 54, (name, total_items)
        assert inv.general_amount is not None and abs(inv.general_amount - 62101.4) < 0.1
        assert inv.general_cartons == 93
        print(f"parser self-check OK ({name})")

    xlsx = root / "NALUYIMA 04 (1).xlsx"
    if xlsx.exists():
        result = parse_file(xlsx)
        assert result.pricing is not None, "expected Sheet2 pricing data"
        assert result.pricing.sheet_name == "Sheet2"
        assert len(result.pricing.items) == 54, len(result.pricing.items)
        assert result.pricing.items[0].unit_cbm == 0.09
        print("parser self-check OK (pricing sheet)")


if __name__ == "__main__":
    _self_check()
